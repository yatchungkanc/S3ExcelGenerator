#!/usr/bin/env python3
from datetime import datetime
import pandas as pd
import json
import os
from openpyxl.worksheet.datavalidation import DataValidation

def load_config(config_file='backup_config.json'):
    """Load configuration from JSON file"""
    if not os.path.exists(config_file):
        raise FileNotFoundError(f"Config file {config_file} not found")
    
    with open(config_file, 'r') as f:
        return json.load(f)

def create_backup_cost_calculator(config_file='backup_config.json'):
    config = load_config(config_file)
    
    pricing = config['pricing']
    min_duration = config['min_duration']
    retrieval_costs = config['retrieval_costs']
    backup_config = config['backup_config']
    storage_size_gb = config['storage_size_gb']
    tier_retention = config['tier_retention']
    aws_backup_pricing = config.get('aws_backup_pricing', {'storage_cost_per_gb_month': 0.05, 'retrieval_cost_per_gb': 0.02})
    
    # Create Excel writer
    filename = f"Backup_Cost_Calculator_{datetime.now().strftime('%Y%m%d')}.xlsx"
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        
        # Create structured configuration layout
        workbook = writer.book
        # Remove default sheet
        if 'Sheet' in workbook.sheetnames:
            workbook.remove(workbook['Sheet'])
        config_sheet = workbook.create_sheet('Configuration')
        
        # Step 1: Backup Tier Selection
        config_sheet['A1'] = 'STEP 1: SELECT BACKUP TIER CONFIGURATION'
        config_sheet['A3'] = 'Backup Tier:'
        config_sheet['B3'] = 'Tier 1'
        
        # Step 2: Backup Configuration
        config_sheet['A6'] = 'STEP 2: BACKUP CONFIGURATION'
        config_sheet['A8'] = 'Backup Type'
        config_sheet['B8'] = 'Count'
        config_sheet['C8'] = 'Storage Tier'
        config_sheet['D8'] = 'Tier Name'
        config_sheet['E8'] = 'Retention (days)'
        
        config_sheet['A9'] = 'Hourly backups'
        config_sheet['B9'] = '=IF(B3="Tier 1",G25,IF(B3="Tier 2",H25,IF(B3="Tier 3",I25,IF(B3="Tier 4",J25,1))))'
        config_sheet['C9'] = backup_config['hourly']['tier']
        config_sheet['D9'] = '=IF(C9=1,"S3 Standard",IF(C9=2,"S3 Intelligent",IF(C9=3,"S3-IA",IF(C9=4,"Glacier IR",IF(C9=5,"Glacier DA","INVALID")))))'
        config_sheet['E9'] = backup_config['hourly']['retention']
        
        config_sheet['A10'] = 'Daily backups'
        config_sheet['B10'] = '=IF(B3="Tier 1",G26,IF(B3="Tier 2",H26,IF(B3="Tier 3",I26,IF(B3="Tier 4",J26,1))))'
        config_sheet['C10'] = backup_config['daily']['tier']
        config_sheet['D10'] = '=IF(C10=1,"S3 Standard",IF(C10=2,"S3 Intelligent",IF(C10=3,"S3-IA",IF(C10=4,"Glacier IR",IF(C10=5,"Glacier DA","INVALID")))))'
        config_sheet['E10'] = backup_config['daily']['retention']
        
        config_sheet['A11'] = 'Weekly backups'
        config_sheet['B11'] = '=IF(B3="Tier 1",G27,IF(B3="Tier 2",H27,IF(B3="Tier 3",I27,IF(B3="Tier 4",J27,1))))'
        config_sheet['C11'] = backup_config['weekly']['tier']
        config_sheet['D11'] = '=IF(C11=1,"S3 Standard",IF(C11=2,"S3 Intelligent",IF(C11=3,"S3-IA",IF(C11=4,"Glacier IR",IF(C11=5,"Glacier DA","INVALID"))))'
        config_sheet['E11'] = tier_retention['S3 Intelligent']
        
        config_sheet['A12'] = 'Off-account backups'
        config_sheet['B12'] = '=IF(B3="Tier 1",G28,IF(B3="Tier 2",H28,IF(B3="Tier 3",I28,IF(B3="Tier 4",J28,1))))'
        config_sheet['C12'] = backup_config['off_account']['tier']
        config_sheet['D12'] = '=IF(C12=1,"S3 Standard",IF(C12=2,"S3 Intelligent",IF(C12=3,"S3-IA",IF(C12=4,"Glacier IR",IF(C12=5,"Glacier DA","INVALID"))))'
        config_sheet['E12'] = tier_retention['Glacier IR']
        
        config_sheet['A14'] = 'Incremental storage size (GB)'
        config_sheet['B14'] = storage_size_gb
        
        # Step 3: Storage Tier Retention
        config_sheet['A16'] = 'STEP 3: STORAGE TIER RETENTION'
        config_sheet['A18'] = 'Storage Tier'
        config_sheet['B18'] = 'Retention (days)'
        config_sheet['C18'] = 'Min Required'
        config_sheet['D18'] = 'Status'
        
        config_sheet['A19'] = 'S3 Standard'
        config_sheet['B19'] = tier_retention['S3 Standard']
        config_sheet['C19'] = min_duration['S3 Standard']
        config_sheet['D19'] = '=IF(B19>=C19,"OK","PENALTY")'
        
        config_sheet['A20'] = 'S3 Intelligent'
        config_sheet['B20'] = tier_retention['S3 Intelligent']
        config_sheet['C20'] = min_duration['S3 Intelligent']
        config_sheet['D20'] = '=IF(B20>=C20,"OK","PENALTY")'
        
        config_sheet['A21'] = 'S3-IA'
        config_sheet['B21'] = tier_retention['S3-IA']
        config_sheet['C21'] = min_duration['S3-IA']
        config_sheet['D21'] = '=IF(B21>=C21,"OK","PENALTY")'
        
        config_sheet['A22'] = 'Glacier IR'
        config_sheet['B22'] = tier_retention['Glacier IR']
        config_sheet['C22'] = min_duration['Glacier IR']
        config_sheet['D22'] = '=IF(B22>=C22,"OK","PENALTY")'
        
        config_sheet['A23'] = 'Glacier Deep Archive'
        config_sheet['B23'] = tier_retention['Glacier Deep Archive']
        config_sheet['C23'] = min_duration['Glacier Deep Archive']
        config_sheet['D23'] = '=IF(B23>=C23,"OK","PENALTY")'
        
        # Backup Tier Reference Table
        config_sheet['F22'] = 'BACKUP TIER REFERENCE'
        config_sheet['F24'] = 'Backup Type'
        config_sheet['G24'] = 'Tier 1'
        config_sheet['H24'] = 'Tier 2'
        config_sheet['I24'] = 'Tier 3'
        config_sheet['J24'] = 'Tier 4'
        
        config_sheet['F25'] = 'Hourly backups'
        config_sheet['G25'] = 24
        config_sheet['H25'] = 24
        config_sheet['I25'] = 6
        config_sheet['J25'] = 0
        
        config_sheet['F26'] = 'Daily backups'
        config_sheet['G26'] = 7
        config_sheet['H26'] = 7
        config_sheet['I26'] = 7
        config_sheet['J26'] = 7
        
        config_sheet['F27'] = 'Weekly backups'
        config_sheet['G27'] = 6
        config_sheet['H27'] = 6
        config_sheet['I27'] = 4
        config_sheet['J27'] = 2
        
        config_sheet['F28'] = 'Off-account backups'
        config_sheet['G28'] = 2
        config_sheet['H28'] = 2
        config_sheet['I28'] = 2
        config_sheet['J28'] = 2
        
        # Format editable cells
        from openpyxl.styles import PatternFill, Font
        blue_fill = PatternFill(start_color='E7F3FF', end_color='E7F3FF', fill_type='solid')
        bold_font = Font(bold=True)
        
        config_sheet['A1'].font = bold_font
        config_sheet['A6'].font = bold_font
        config_sheet['A16'].font = bold_font
        config_sheet['F22'].font = bold_font
        
        # Add dropdown for backup tier selection
        dv = DataValidation(type="list", formula1='"Tier 1,Tier 2,Tier 3,Tier 4"')
        config_sheet.add_data_validation(dv)
        dv.add('B3')
        config_sheet['B3'].fill = blue_fill
        
        # Make all Count, Storage Tier, and Retention cells editable with blue fill
        for row in [9, 10, 11, 12]:  # All backup types
            config_sheet[f'B{row}'].fill = blue_fill  # Count
            config_sheet[f'C{row}'].fill = blue_fill  # Storage Tier
            config_sheet[f'E{row}'].fill = blue_fill  # Retention
        config_sheet['B14'].fill = blue_fill  # Storage size
        
        # Make retention days editable
        for row in range(19, 24):
            config_sheet[f'B{row}'].fill = blue_fill
        
        # Pricing sheet
        pricing_data = {
            'Storage Tier': list(pricing.keys()) + ['AWS Backup'],
            'Price per GB/month (USD)': list(pricing.values()) + [aws_backup_pricing['storage_cost_per_gb_month']],
            'Min Duration (days)': [min_duration[tier] for tier in pricing.keys()] + [0],
            'Retrieval Cost per GB (USD)': [retrieval_costs[tier] for tier in pricing.keys()] + [aws_backup_pricing['retrieval_cost_per_gb']]
        }
        
        pricing_df = pd.DataFrame(pricing_data)
        pricing_df.to_excel(writer, sheet_name='Pricing', index=False)
        
        # Create calculation sheet with formulas
        calc_data = {
            'Month': [f'Month {i}' for i in range(1, 13)],
            'Total Backups (GB)': [''] * 12,
            'S3 Standard (GB)': [''] * 12,
            'S3 Intelligent (GB)': [''] * 12,
            'S3-IA (GB)': [''] * 12,
            'Glacier IR (GB)': [''] * 12,
            'Glacier DA (GB)': [''] * 12,
            'Total Storage (GB)': [''] * 12,
            'Storage Cost (USD)': [''] * 12,
            'Early Deletion Penalty (USD)': [''] * 12,
            'Total Monthly Cost (USD)': [''] * 12,
            'AWS Backup Cost (USD)': [''] * 12,
            'Cost Difference (USD)': [''] * 12
        }
        
        calc_df = pd.DataFrame(calc_data)
        calc_df.to_excel(writer, sheet_name='Cost Calculation', index=False)
        
        # Add formulas to the calculation sheet
        workbook = writer.book
        calc_sheet = workbook['Cost Calculation']
        
        for row in range(2, 14):  # Rows 2-13 (months 1-12)
            month = row - 1
            
            # Total backups in GB
            calc_sheet[f'B{row}'] = f'=(Configuration!B9+Configuration!B10+Configuration!B11+Configuration!B12)*Configuration!B14'
            
            # S3 Standard storage
            calc_sheet[f'C{row}'] = f'=(IF(Configuration!C9=1,Configuration!B9*Configuration!B14*MIN(Configuration!E9,{month}*30)/30,0)+IF(Configuration!C10=1,Configuration!B10*Configuration!B14*MIN(Configuration!E10,{month}*30)/30,0)+IF(Configuration!C11=1,Configuration!B11*Configuration!B14*MIN(Configuration!B19,{month}*30)/30,0)+IF(Configuration!C12=1,Configuration!B12*Configuration!B14*MIN(Configuration!B19,{month}*30)/30,0))'
            
            # S3 Intelligent storage
            calc_sheet[f'D{row}'] = f'=IF(Configuration!B20>0,(IF(Configuration!C9=2,Configuration!B9,0)+IF(Configuration!C10=2,Configuration!B10,0)+IF(Configuration!C11=2,Configuration!B11,0)+IF(Configuration!C12=2,Configuration!B12,0))*Configuration!B14*MIN(Configuration!B20,{month}*30)/30,0)'
            
            # S3-IA storage
            calc_sheet[f'E{row}'] = f'=IF(Configuration!B21>0,(IF(Configuration!C9=3,Configuration!B9,0)+IF(Configuration!C10=3,Configuration!B10,0)+IF(Configuration!C11=3,Configuration!B11,0)+IF(Configuration!C12=3,Configuration!B12,0))*Configuration!B14*MIN(Configuration!B21,{month}*30)/30,0)'
            
            # Glacier IR storage
            calc_sheet[f'F{row}'] = f'=IF(Configuration!B22>0,(IF(Configuration!C9=4,Configuration!B9,0)+IF(Configuration!C10=4,Configuration!B10,0)+IF(Configuration!C11=4,Configuration!B11,0)+IF(Configuration!C12=4,Configuration!B12,0))*Configuration!B14*MIN(Configuration!B22,{month}*30)/30,0)'
            
            # Glacier Deep Archive storage
            calc_sheet[f'G{row}'] = f'=IF(Configuration!B23>0,(IF(Configuration!C9=5,Configuration!B9,0)+IF(Configuration!C10=5,Configuration!B10,0)+IF(Configuration!C11=5,Configuration!B11,0)+IF(Configuration!C12=5,Configuration!B12,0))*Configuration!B14*MIN(Configuration!B23,{month}*30)/30,0)'
            
            # Total storage
            calc_sheet[f'H{row}'] = f'=C{row}+D{row}+E{row}+F{row}+G{row}'
            
            # Storage cost
            calc_sheet[f'I{row}'] = f'=(IF(Configuration!C9=1,Configuration!B9*Configuration!B14*MIN(Configuration!E9,{month}*30)/30*Pricing!B2*(Configuration!E9/30),0)+IF(Configuration!C10=1,Configuration!B10*Configuration!B14*MIN(Configuration!E10,{month}*30)/30*Pricing!B2*(Configuration!E10/30),0)+IF(Configuration!C11=1,Configuration!B11*Configuration!B14*MIN(Configuration!B19,{month}*30)/30*Pricing!B2*(Configuration!B19/30),0)+IF(Configuration!C12=1,Configuration!B12*Configuration!B14*MIN(Configuration!B19,{month}*30)/30*Pricing!B2*(Configuration!B19/30),0))+D{row}*Pricing!B3*(Configuration!B20/30)+E{row}*Pricing!B4*(Configuration!B21/30)+F{row}*Pricing!B5*(Configuration!B22/30)+G{row}*Pricing!B6*(Configuration!B23/30)'
            
            # Early deletion penalty
            calc_sheet[f'J{row}'] = f'=IF(Configuration!B21<Pricing!C4,E{row}*Pricing!B4*((Pricing!C4-Configuration!B21)/30),0)+IF(Configuration!B22<Pricing!C5,F{row}*Pricing!B5*((Pricing!C5-Configuration!B22)/30),0)+IF(Configuration!B23<Pricing!C6,G{row}*Pricing!B6*((Pricing!C6-Configuration!B23)/30),0)'
            
            # Total monthly cost
            calc_sheet[f'K{row}'] = f'=I{row}+J{row}'
            
            # AWS Backup cost
            calc_sheet[f'L{row}'] = f'=H{row}*Pricing!B7'
            
            # Cost difference
            calc_sheet[f'M{row}'] = f'=K{row}-L{row}'
        
        # Add annual summary
        calc_sheet['A14'] = 'Annual Total'
        for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M']:
            calc_sheet[f'{col}14'] = f'=SUM({col}2:{col}13)'
        
        # Format pricing sheet values as editable
        pricing_sheet = workbook['Pricing']
        for row in range(2, 8):
            pricing_sheet[f'B{row}'].fill = blue_fill
        
        # Remove default sheet to prevent corruption
        default_sheets = ['Sheet', 'Sheet1', 'Worksheet']
        for sheet_name in default_sheets:
            if sheet_name in workbook.sheetnames:
                workbook.remove(workbook[sheet_name])
        
        # Auto-adjust column widths
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            for column in sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 20)
                sheet.column_dimensions[column_letter].width = adjusted_width
    
    print(f"Backup cost calculator created: {filename}")
    return filename

if __name__ == "__main__":
    create_backup_cost_calculator()