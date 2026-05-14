#!/usr/bin/env python3
from datetime import datetime
import json
import os
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import PatternFill, Font

def load_config(config_file='backup_config.json'):
    """Load configuration from JSON file"""
    if not os.path.exists(config_file):
        raise FileNotFoundError(f"Config file {config_file} not found")
    
    with open(config_file, 'r') as f:
        return json.load(f)

def create_backup_cost_calculator(config_file='backup_config.json'):
    config = load_config(config_file)
    
    pricing = config['pricing']
    min_duration = config['min_duration']
    retrieval_costs = config['retrieval_costs']
    backup_config = config['backup_config']
    storage_size_gb = config['storage_size_gb']
    tier_retention = config['tier_retention']
    aws_backup_pricing = config.get('aws_backup_pricing', {'storage_cost_per_gb_month': 0.05, 'retrieval_cost_per_gb': 0.02})
    
    # Create workbook directly
    workbook = Workbook()
    workbook.remove(workbook.active)  # Remove default sheet
    
    # Configuration sheet
    config_sheet = workbook.create_sheet('Configuration')
    
    # Step 1: Backup Tier Selection
    config_sheet['A1'] = 'STEP 1: SELECT BACKUP TIER CONFIGURATION'
    config_sheet['A3'] = 'Backup Tier:'
    config_sheet['B3'] = 'Tier 1'
    
    # Step 2: Backup Configuration
    config_sheet['A6'] = 'STEP 2: BACKUP CONFIGURATION'
    config_sheet['A8'] = 'Backup Type'
    config_sheet['B8'] = 'Count'
    config_sheet['C8'] = 'Storage Tier'
    config_sheet['D8'] = 'Tier Name'
    config_sheet['E8'] = 'Retention (days)'
    
    config_sheet['A9'] = 'Hourly backups'
    config_sheet['B9'] = '=IF(B3="Tier 1",G25,IF(B3="Tier 2",H25,IF(B3="Tier 3",I25,IF(B3="Tier 4",J25,1))))'
    config_sheet['C9'] = backup_config['hourly']['tier']
    config_sheet['D9'] = '=IF(C9=1,"S3 Standard",IF(C9=2,"S3 Intelligent",IF(C9=3,"S3-IA",IF(C9=4,"Glacier IR",IF(C9=5,"Glacier DA","INVALID")))))'
    config_sheet['E9'] = backup_config['hourly']['retention']
    
    config_sheet['A10'] = 'Daily backups'
    config_sheet['B10'] = '=IF(B3="Tier 1",G26,IF(B3="Tier 2",H26,IF(B3="Tier 3",I26,IF(B3="Tier 4",J26,1))))'
    config_sheet['C10'] = backup_config['daily']['tier']
    config_sheet['D10'] = '=IF(C10=1,"S3 Standard",IF(C10=2,"S3 Intelligent",IF(C10=3,"S3-IA",IF(C10=4,"Glacier IR",IF(C10=5,"Glacier DA","INVALID")))))'
    config_sheet['E10'] = backup_config['daily']['retention']
    
    config_sheet['A11'] = 'Weekly backups'
    config_sheet['B11'] = '=IF(B3="Tier 1",G27,IF(B3="Tier 2",H27,IF(B3="Tier 3",I27,IF(B3="Tier 4",J27,1))))'
    config_sheet['C11'] = backup_config['weekly']['tier']
    config_sheet['D11'] = '=IF(C11=1,"S3 Standard",IF(C11=2,"S3 Intelligent",IF(C11=3,"S3-IA",IF(C11=4,"Glacier IR",IF(C11=5,"Glacier DA","INVALID")))))'
    config_sheet['E11'] = tier_retention['S3 Intelligent']
    
    config_sheet['A12'] = 'Off-account backups'
    config_sheet['B12'] = '=IF(B3="Tier 1",G28,IF(B3="Tier 2",H28,IF(B3="Tier 3",I28,IF(B3="Tier 4",J28,1))))'
    config_sheet['C12'] = backup_config['off_account']['tier']
    config_sheet['D12'] = '=IF(C12=1,"S3 Standard",IF(C12=2,"S3 Intelligent",IF(C12=3,"S3-IA",IF(C12=4,"Glacier IR",IF(C12=5,"Glacier DA","INVALID")))))'
    config_sheet['E12'] = tier_retention['Glacier IR']
    
    config_sheet['A14'] = 'Incremental storage size (GB)'
    config_sheet['B14'] = storage_size_gb
    
    # Step 3: Storage Tier Retention
    config_sheet['A16'] = 'STEP 3: STORAGE TIER RETENTION'
    config_sheet['A17'] = 'Note: These values are used for storage calculations when backup types do not have individual retention periods'
    config_sheet['A19'] = 'Storage Tier'
    config_sheet['B19'] = 'Retention (days)'
    config_sheet['C19'] = 'Min Required'
    config_sheet['D19'] = 'Status'
    
    config_sheet['A20'] = 'S3 Standard'
    config_sheet['B20'] = tier_retention['S3 Standard']
    config_sheet['C20'] = min_duration['S3 Standard']
    config_sheet['D20'] = '=IF(B20>=C20,"OK","PENALTY")'
    
    config_sheet['A21'] = 'S3 Intelligent'
    config_sheet['B21'] = tier_retention['S3 Intelligent']
    config_sheet['C21'] = min_duration['S3 Intelligent']
    config_sheet['D21'] = '=IF(B21>=C21,"OK","PENALTY")'
    
    config_sheet['A22'] = 'S3-IA'
    config_sheet['B22'] = tier_retention['S3-IA']
    config_sheet['C22'] = min_duration['S3-IA']
    config_sheet['D22'] = '=IF(B22>=C22,"OK","PENALTY")'
    
    config_sheet['A23'] = 'Glacier IR'
    config_sheet['B23'] = tier_retention['Glacier IR']
    config_sheet['C23'] = min_duration['Glacier IR']
    config_sheet['D23'] = '=IF(B23>=C23,"OK","PENALTY")'
    
    config_sheet['A24'] = 'Glacier Deep Archive'
    config_sheet['B24'] = tier_retention['Glacier Deep Archive']
    config_sheet['C24'] = min_duration['Glacier Deep Archive']
    config_sheet['D24'] = '=IF(B24>=C24,"OK","PENALTY")'
    
    # Backup Tier Reference Table
    config_sheet['F22'] = 'BACKUP TIER REFERENCE'
    config_sheet['F24'] = 'Backup Type'
    config_sheet['G24'] = 'Tier 1'
    config_sheet['H24'] = 'Tier 2'
    config_sheet['I24'] = 'Tier 3'
    config_sheet['J24'] = 'Tier 4'
    
    config_sheet['F25'] = 'Hourly backups'
    config_sheet['G25'] = 24
    config_sheet['H25'] = 24
    config_sheet['I25'] = 6
    config_sheet['J25'] = 0
    
    config_sheet['F26'] = 'Daily backups'
    config_sheet['G26'] = 7
    config_sheet['H26'] = 7
    config_sheet['I26'] = 7
    config_sheet['J26'] = 7
    
    config_sheet['F27'] = 'Weekly backups'
    config_sheet['G27'] = 6
    config_sheet['H27'] = 6
    config_sheet['I27'] = 4
    config_sheet['J27'] = 2
    
    config_sheet['F28'] = 'Off-account backups'
    config_sheet['G28'] = 2
    config_sheet['H28'] = 2
    config_sheet['I28'] = 2
    config_sheet['J28'] = 2
    
    # Format cells
    blue_fill = PatternFill(start_color='E7F3FF', end_color='E7F3FF', fill_type='solid')
    bold_font = Font(bold=True)
    
    config_sheet['A1'].font = bold_font
    config_sheet['A6'].font = bold_font
    config_sheet['A16'].font = bold_font
    config_sheet['F22'].font = bold_font
    
    # Add dropdown
    dv = DataValidation(type="list", formula1='"Tier 1,Tier 2,Tier 3,Tier 4"')
    config_sheet.add_data_validation(dv)
    dv.add('B3')
    config_sheet['B3'].fill = blue_fill
    
    # Make editable cells blue
    for row in [9, 10, 11, 12]:
        config_sheet[f'B{row}'].fill = blue_fill
        config_sheet[f'C{row}'].fill = blue_fill
        config_sheet[f'E{row}'].fill = blue_fill
    config_sheet['B14'].fill = blue_fill
    
    for row in range(20, 25):
        config_sheet[f'B{row}'].fill = blue_fill
    
    # Pricing sheet
    pricing_sheet = workbook.create_sheet('Pricing')
    pricing_sheet['A1'] = 'Storage Tier'
    pricing_sheet['B1'] = 'Price per GB/month (USD)'
    pricing_sheet['C1'] = 'Min Duration (days)'
    pricing_sheet['D1'] = 'Retrieval Cost per GB (USD)'
    
    # Format headers
    for col in ['A1', 'B1', 'C1', 'D1']:
        pricing_sheet[col].font = bold_font
    
    row = 2
    for tier in pricing.keys():
        pricing_sheet[f'A{row}'] = tier
        pricing_sheet[f'B{row}'] = pricing[tier]
        pricing_sheet[f'C{row}'] = min_duration[tier]
        pricing_sheet[f'D{row}'] = retrieval_costs[tier]
        pricing_sheet[f'B{row}'].fill = blue_fill
        row += 1
    
    pricing_sheet[f'A{row}'] = 'AWS Backup'
    pricing_sheet[f'B{row}'] = aws_backup_pricing['storage_cost_per_gb_month']
    pricing_sheet[f'C{row}'] = 0
    pricing_sheet[f'D{row}'] = aws_backup_pricing['retrieval_cost_per_gb']
    pricing_sheet[f'B{row}'].fill = blue_fill
    
    # Auto-adjust column widths for pricing
    for column in pricing_sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 20)
        pricing_sheet.column_dimensions[column_letter].width = adjusted_width
    
    # Cost Calculation sheet
    calc_sheet = workbook.create_sheet('Cost Calculation')
    headers = ['Month', 'Total Backups (GB)', 'S3 Standard (GB)', 'S3 Intelligent (GB)', 
               'S3-IA (GB)', 'Glacier IR (GB)', 'Glacier DA (GB)', 'Total Storage (GB)',
               'Storage Cost (USD)', 'Early Deletion Penalty (USD)', 'Total Monthly Cost (USD)',
               'AWS Backup Cost (USD)', 'Cost Difference (USD)']
    
    for col, header in enumerate(headers, 1):
        cell = calc_sheet.cell(row=1, column=col, value=header)
        cell.font = bold_font
    
    for month in range(1, 13):
        row = month + 1
        calc_sheet[f'A{row}'] = f'Month {month}'
        calc_sheet[f'B{row}'] = f'=(Configuration!B9+Configuration!B10+Configuration!B11+Configuration!B12)*Configuration!B14'
        calc_sheet[f'C{row}'] = f'=(IF(Configuration!C9=1,Configuration!B9*Configuration!B14*MIN(Configuration!E9,{month}*30)/30,0)+IF(Configuration!C10=1,Configuration!B10*Configuration!B14*MIN(Configuration!E10,{month}*30)/30,0)+IF(Configuration!C11=1,Configuration!B11*Configuration!B14*MIN(Configuration!E11,{month}*30)/30,0)+IF(Configuration!C12=1,Configuration!B12*Configuration!B14*MIN(Configuration!E12,{month}*30)/30,0))'
        calc_sheet[f'D{row}'] = f'=IF(Configuration!B21>0,(IF(Configuration!C9=2,Configuration!B9*MIN(Configuration!E9,{month}*30)/30,0)+IF(Configuration!C10=2,Configuration!B10*MIN(Configuration!E10,{month}*30)/30,0)+IF(Configuration!C11=2,Configuration!B11*MIN(Configuration!E11,{month}*30)/30,0)+IF(Configuration!C12=2,Configuration!B12*MIN(Configuration!E12,{month}*30)/30,0))*Configuration!B14,0)'
        calc_sheet[f'E{row}'] = f'=IF(Configuration!B22>0,(IF(Configuration!C9=3,Configuration!B9*MIN(Configuration!E9,{month}*30)/30,0)+IF(Configuration!C10=3,Configuration!B10*MIN(Configuration!E10,{month}*30)/30,0)+IF(Configuration!C11=3,Configuration!B11*MIN(Configuration!E11,{month}*30)/30,0)+IF(Configuration!C12=3,Configuration!B12*MIN(Configuration!E12,{month}*30)/30,0))*Configuration!B14,0)'
        calc_sheet[f'F{row}'] = f'=IF(Configuration!B23>0,(IF(Configuration!C9=4,Configuration!B9*MIN(Configuration!E9,{month}*30)/30,0)+IF(Configuration!C10=4,Configuration!B10*MIN(Configuration!E10,{month}*30)/30,0)+IF(Configuration!C11=4,Configuration!B11*MIN(Configuration!E11,{month}*30)/30,0)+IF(Configuration!C12=4,Configuration!B12*MIN(Configuration!E12,{month}*30)/30,0))*Configuration!B14,0)'
        calc_sheet[f'G{row}'] = f'=IF(Configuration!B24>0,(IF(Configuration!C9=5,Configuration!B9*MIN(Configuration!E9,{month}*30)/30,0)+IF(Configuration!C10=5,Configuration!B10*MIN(Configuration!E10,{month}*30)/30,0)+IF(Configuration!C11=5,Configuration!B11*MIN(Configuration!E11,{month}*30)/30,0)+IF(Configuration!C12=5,Configuration!B12*MIN(Configuration!E12,{month}*30)/30,0))*Configuration!B14,0)'
        calc_sheet[f'H{row}'] = f'=C{row}+D{row}+E{row}+F{row}+G{row}'
        calc_sheet[f'I{row}'] = f'=(IF(Configuration!C9=1,Configuration!B9*Configuration!B14*MIN(Configuration!E9,{month}*30)/30*Pricing!B2,0)+IF(Configuration!C10=1,Configuration!B10*Configuration!B14*MIN(Configuration!E10,{month}*30)/30*Pricing!B2,0)+IF(Configuration!C11=1,Configuration!B11*Configuration!B14*MIN(Configuration!E11,{month}*30)/30*Pricing!B2,0)+IF(Configuration!C12=1,Configuration!B12*Configuration!B14*MIN(Configuration!E12,{month}*30)/30*Pricing!B2,0))+D{row}*Pricing!B3+E{row}*Pricing!B4+F{row}*Pricing!B5+G{row}*Pricing!B6'
        calc_sheet[f'J{row}'] = f'=(IF(Configuration!C9=3,IF(Configuration!E9<Pricing!C4,Configuration!B9*Configuration!B14*MIN(Configuration!E9,{month}*30)/30*Pricing!B4*((Pricing!C4-Configuration!E9)/30),0),0)+IF(Configuration!C10=3,IF(Configuration!E10<Pricing!C4,Configuration!B10*Configuration!B14*MIN(Configuration!E10,{month}*30)/30*Pricing!B4*((Pricing!C4-Configuration!E10)/30),0),0)+IF(Configuration!C11=3,IF(Configuration!E11<Pricing!C4,Configuration!B11*Configuration!B14*MIN(Configuration!E11,{month}*30)/30*Pricing!B4*((Pricing!C4-Configuration!E11)/30),0),0)+IF(Configuration!C12=3,IF(Configuration!E12<Pricing!C4,Configuration!B12*Configuration!B14*MIN(Configuration!E12,{month}*30)/30*Pricing!B4*((Pricing!C4-Configuration!E12)/30),0),0)+IF(Configuration!C9=4,IF(Configuration!E9<Pricing!C5,Configuration!B9*Configuration!B14*MIN(Configuration!E9,{month}*30)/30*Pricing!B5*((Pricing!C5-Configuration!E9)/30),0),0)+IF(Configuration!C10=4,IF(Configuration!E10<Pricing!C5,Configuration!B10*Configuration!B14*MIN(Configuration!E10,{month}*30)/30*Pricing!B5*((Pricing!C5-Configuration!E10)/30),0),0)+IF(Configuration!C11=4,IF(Configuration!E11<Pricing!C5,Configuration!B11*Configuration!B14*MIN(Configuration!E11,{month}*30)/30*Pricing!B5*((Pricing!C5-Configuration!E11)/30),0),0)+IF(Configuration!C12=4,IF(Configuration!E12<Pricing!C5,Configuration!B12*Configuration!B14*MIN(Configuration!E12,{month}*30)/30*Pricing!B5*((Pricing!C5-Configuration!E12)/30),0),0)+IF(Configuration!C9=5,IF(Configuration!E9<Pricing!C6,Configuration!B9*Configuration!B14*MIN(Configuration!E9,{month}*30)/30*Pricing!B6*((Pricing!C6-Configuration!E9)/30),0),0)+IF(Configuration!C10=5,IF(Configuration!E10<Pricing!C6,Configuration!B10*Configuration!B14*MIN(Configuration!E10,{month}*30)/30*Pricing!B6*((Pricing!C6-Configuration!E10)/30),0),0)+IF(Configuration!C11=5,IF(Configuration!E11<Pricing!C6,Configuration!B11*Configuration!B14*MIN(Configuration!E11,{month}*30)/30*Pricing!B6*((Pricing!C6-Configuration!E11)/30),0),0)+IF(Configuration!C12=5,IF(Configuration!E12<Pricing!C6,Configuration!B12*Configuration!B14*MIN(Configuration!E12,{month}*30)/30*Pricing!B6*((Pricing!C6-Configuration!E12)/30),0),0))'
        calc_sheet[f'K{row}'] = f'=I{row}+J{row}'
        calc_sheet[f'L{row}'] = f'=H{row}*Pricing!B7'
        calc_sheet[f'M{row}'] = f'=K{row}-L{row}'
    
    # Annual totals
    calc_sheet['A14'] = 'Annual Total'
    calc_sheet['A14'].font = bold_font
    for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M']:
        calc_sheet[f'{col}14'] = f'=SUM({col}2:{col}13)'
    
    # Auto-adjust column widths for calculation sheet
    for column in calc_sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 20)
        calc_sheet.column_dimensions[column_letter].width = adjusted_width
    
    # Auto-adjust column widths for configuration sheet
    for column in config_sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 20)
        config_sheet.column_dimensions[column_letter].width = adjusted_width
    
    # Save file
    filename = f"Backup_Cost_Calculator_{datetime.now().strftime('%Y%m%d')}.xlsx"
    workbook.save(filename)
    
    print(f"Backup cost calculator created: {filename}")
    return filename

if __name__ == "__main__":
    create_backup_cost_calculator()
