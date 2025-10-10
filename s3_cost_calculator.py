#!/usr/bin/env python3
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import datetime

def create_s3_cost_calculator():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "S3 Cost Calculator"
    
    # Header styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    input_fill = PatternFill(start_color="E7F3FF", end_color="E7F3FF", fill_type="solid")
    
    # Title
    ws['A1'] = "S3 Storage Cost Calculator (1 Year)"
    ws['A1'].font = Font(bold=True, size=16)
    ws.merge_cells('A1:H1')
    
    # Input section
    ws['A3'] = "Input Parameters"
    ws['A3'].font = header_font
    ws['A3'].fill = header_fill
    ws.merge_cells('A3:B3')
    
    # Input fields
    inputs = [
        ("Base Size (GB)", "B4"),
        ("Daily Incremental Size (GB)", "B5"),
        ("Storage Period (months)", "B6"),
        ("Delete Standard data after (days)", "B7"),
        ("Transition to IA after (days)", "B8"),
        ("Transition to Glacier IR after (days)", "B9"),
        ("Delete Glacier IR data after (days)", "B10"),
        ("Glacier IR min storage period (days)", "B11")
    ]
    
    for i, (label, cell) in enumerate(inputs, 4):
        ws[f'A{i}'] = label
        ws[cell].fill = input_fill
        ws[cell] = 100 if i == 4 else (2 if i == 5 else (12 if i == 6 else (15 if i == 7 else (30 if i == 8 else (90 if i == 9 else (180 if i == 10 else 90))))))
    
    # Pricing section
    ws['A13'] = "S3 Pricing (USD per GB/month)"
    ws['A13'].font = header_font
    ws['A13'].fill = header_fill
    ws.merge_cells('A13:C13')
    
    pricing = [
        ("Standard", "B14", 0.023),
        ("Standard-IA", "B15", 0.0125),
        ("Glacier Instant Retrieval", "B16", 0.004),
        ("Glacier IR Early Deletion Fee", "B17", 0.004)
    ]
    
    for label, cell, price in pricing:
        row = int(cell[1:])
        ws[f'A{row}'] = label
        ws[cell] = price
        ws[cell].fill = input_fill
    
    # Calculation table headers
    ws['A19'] = "Month"
    ws['B19'] = "Total Size (GB)"
    ws['C19'] = "Standard (GB)"
    ws['D19'] = "Standard-IA (GB)"
    ws['E19'] = "Glacier IR (GB)"
    ws['F19'] = "Deleted (GB)"
    ws['G19'] = "Standard Cost"
    ws['H19'] = "IA Cost"
    ws['I19'] = "Glacier IR Cost"
    ws['J19'] = "Glacier IR Del Fee"
    ws['K19'] = "Monthly Total"
    
    for col in range(1, 12):
        cell = ws.cell(row=19, column=col)
        cell.font = header_font
        cell.fill = header_fill
    
    # Monthly calculations
    for month in range(1, 13):
        row = 19 + month
        ws[f'A{row}'] = month
        
        # Total size calculation (daily incremental)
        ws[f'B{row}'] = f'=$B$4+($B$5*({month-1}*30))'
        
        # Storage tier calculations on total storage
        ws[f'C{row}'] = f'=MIN(B{row},$B$5*$B$8)'
        ws[f'D{row}'] = f'=MAX(0,MIN(B{row}-C{row},$B$5*($B$9-$B$8)))'
        ws[f'E{row}'] = f'=MAX(0,B{row}-C{row}-D{row})'
        
        # Deleted data calculation (only Glacier IR deletion)
        ws[f'F{row}'] = f'=IF({month}*30>$B$10,MAX(0,$B$5*({month}*30-$B$10)),0)'
        
        # Cost calculations with Standard prorated for early deletion
        # Standard: prorated if deleted before transition to IA
        standard_days = f'IF({month}*30>$B$7,MIN($B$7,30),30)'
        ws[f'G{row}'] = f'=C{row}*$B$14*({standard_days}/30)'
        ws[f'H{row}'] = f'=D{row}*$B$15'
        ws[f'I{row}'] = f'=E{row}*$B$16'
        
        # Glacier IR early deletion fee only
        ws[f'J{row}'] = f'=IF(F{row}>0,F{row}*$B$17*MAX(0,($B$11-({month}*30-$B$9))/30),0)'
        
        ws[f'K{row}'] = f'=G{row}+H{row}+I{row}+J{row}'
    
    # Summary section
    ws['A33'] = "Summary"
    ws['A33'].font = header_font
    ws['A33'].fill = header_fill
    ws.merge_cells('A33:B33')
    
    ws['A34'] = "Total Annual Cost:"
    ws['B34'] = "=SUM(K20:K31)"
    ws['B34'].font = Font(bold=True)
    
    ws['A35'] = "Glacier IR Deletion Fees:"
    ws['B35'] = "=SUM(J20:J31)"
    
    ws['A36'] = "Average Monthly Cost:"
    ws['B36'] = "=B34/12"
    
    ws['A37'] = "Final Storage Size:"
    ws['B37'] = "=B31"
    
    # Format columns
    for col in range(2, 12):
        for row in range(20, 32):
            ws.cell(row=row, column=col).number_format = '#,##0.00'
    
    ws['B34'].number_format = '$#,##0.00'
    ws['B35'].number_format = '$#,##0.00'
    ws['B36'].number_format = '$#,##0.00'
    ws['B37'].number_format = '#,##0.00'
    
    # Auto-adjust column widths
    for col in range(1, 12):
        ws.column_dimensions[get_column_letter(col)].width = 15
    
    # Add instructions
    ws['A39'] = "Instructions:"
    ws['A39'].font = Font(bold=True)
    ws['A40'] = "1. Modify blue cells to adjust parameters"
    ws['A41'] = "2. Standard cost is prorated based on storage duration"
    ws['A42'] = "3. Glacier IR min storage: 90 days (early deletion fees apply)"
    ws['A43'] = "4. No early deletion fees for Standard storage"
    
    return wb

if __name__ == "__main__":
    workbook = create_s3_cost_calculator()
    filename = f"S3_Cost_Calculator_{datetime.datetime.now().strftime('%Y%m%d')}.xlsx"
    workbook.save(filename)
    print(f"Excel file created: {filename}")